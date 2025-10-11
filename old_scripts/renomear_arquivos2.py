import os
from openpyxl import load_workbook
import re
from datetime import datetime

def limpar_nome(nome):
    """Remove caracteres inválidos do Windows."""
    return re.sub(r'[<>:"/\\|?*]', '', nome)

def run(log_callback=None, pasta_videos="Karaoke", arquivo_xlsx="karaoke.xlsx"):
    if not os.path.exists(pasta_videos):
        if log_callback:
            log_callback(f"Pasta não encontrada, criando: {pasta_videos}")
        os.makedirs(pasta_videos, exist_ok=True)

    if not os.path.exists(arquivo_xlsx):
        if log_callback:
            log_callback(f"Arquivo XLSX não encontrado: {arquivo_xlsx}")
        return

    try:
        wb = load_workbook(arquivo_xlsx)
    except Exception as e:
        if log_callback:
            log_callback(f"Erro ao abrir o arquivo XLSX: {e}")
        return

    if "Normaliza" not in wb.sheetnames:
        if log_callback:
            log_callback("Erro: A aba 'Normaliza' não foi encontrada no arquivo Excel.")
        return

    ws = wb["Normaliza"]
    header = [cell.value for cell in ws[1]]

    try:
        idx_filename = header.index("filename") + 1
        idx_fulltitle = header.index("full_title") + 1
    except ValueError as e:
        if log_callback:
            log_callback(f"Erro: Colunas 'filename' ou 'full_title' não encontradas. {e}")
        return

    renomeacoes = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        filename = row[idx_filename - 1]
        fulltitle = row[idx_fulltitle - 1]

        if not filename:
            if log_callback:
                log_callback(f"Linha {i}: Sem filename, ignorando...")
            continue
        if not fulltitle:
            if log_callback:
                log_callback(f"Linha {i}: Sem full_title, ignorando...")
            continue

        # Percorre todas as subpastas
        encontrado = False
        for root, dirs, files in os.walk(pasta_videos):
            if filename in files:
                caminho_atual = os.path.join(root, filename)
                ext = os.path.splitext(filename)[1]
                novo_nome = limpar_nome(fulltitle) + ext
                caminho_novo = os.path.join(root, novo_nome)

                contador = 2
                while os.path.exists(caminho_novo):
                    novo_nome = limpar_nome(fulltitle) + f" v{contador}" + ext
                    caminho_novo = os.path.join(root, novo_nome)
                    contador += 1

                try:
                    os.rename(caminho_atual, caminho_novo)
                    if log_callback:
                        log_callback(f"Linha {i}: {filename} -> {novo_nome}")
                    renomeacoes.append(f"{caminho_atual} => {caminho_novo}")
                    encontrado = True
                except Exception as e:
                    if log_callback:
                        log_callback(f"Erro na linha {i}: {e}")
                break

        if not encontrado and log_callback:
            log_callback(f"Linha {i}: Arquivo não encontrado: {filename}")

    # Gera log final
    if renomeacoes:
        agora = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_dir = os.path.join(os.path.dirname(__file__), "..", "logs")
        os.makedirs(log_dir, exist_ok=True)
        log_filename = os.path.join(log_dir, f"karaoke_renomear_{agora}.log")
        with open(log_filename, "w", encoding="utf-8") as log_file:
            log_file.write("\n".join(renomeacoes))
        if log_callback:
            log_callback(f"\nLog gerado: {log_filename}")
    else:
        if log_callback:
            log_callback("\nNenhum arquivo renomeado. Log não gerado.")

# Permite execução standalone
if __name__ == "__main__":
    run()
