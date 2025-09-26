import os
from openpyxl import load_workbook
import re
from datetime import datetime

def limpar_nome(nome):
    # Remove caracteres inválidos do Windows
    return re.sub(r'[<>:"/\\|?*]', '', nome)

def renomear_arquivos():
    pasta = "Karaoke"
    arquivo_excel = "karaoke.xlsx"

    wb = load_workbook(arquivo_excel)

    # Garante que está pegando a aba correta
    if "Normaliza" not in wb.sheetnames:
        print("Erro: A aba 'Normaliza' não foi encontrada no arquivo Excel.")
        return

    ws = wb["Normaliza"]

    header = [cell.value for cell in ws[1]]
    idx_filename = header.index("filename") + 1
    idx_fulltitle = header.index("full_title") + 1

    # Lista para registrar os arquivos renomeados
    renomeacoes = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        filename = row[idx_filename - 1]
        fulltitle = row[idx_fulltitle - 1]

        # Se não houver filename, ignora a linha sem gastar tempo
        if not filename:
            print(f"Linha {i}: Sem filename, ignorando...")
            continue

        # Se não houver título, também não faz sentido renomear
        if not fulltitle:
            print(f"Linha {i}: Sem full_title, ignorando...")
            continue

        base, ext = os.path.splitext(filename)
        novo_nome = limpar_nome(fulltitle) + ext
        caminho_atual = os.path.join(pasta, filename)
        caminho_novo = os.path.join(pasta, novo_nome)

        try:
            if os.path.exists(caminho_atual):
                contador = 2
                while os.path.exists(caminho_novo):
                    novo_nome = limpar_nome(fulltitle) + f" v{contador}" + ext
                    caminho_novo = os.path.join(pasta, novo_nome)
                    contador += 1

                os.rename(caminho_atual, caminho_novo)
                print(f"Linha {i}: {filename} -> {novo_nome}")

                # Adiciona ao log
                renomeacoes.append(f"{filename} => {novo_nome}")
            else:
                print(f"Linha {i}: Arquivo não encontrado: {filename}")
        except Exception as e:
            print(f"Erro na linha {i}: {e}")

    # Gera o log apenas se houver renomeações
    if renomeacoes:
        agora = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_filename = f"karaoke_renomear_{agora}.log"
        with open(log_filename, "w", encoding="utf-8") as log_file:
            log_file.write("\n".join(renomeacoes))
        print(f"\nLog gerado: {log_filename}")
    else:
        print("\nNenhum arquivo renomeado. Log não gerado.")

if __name__ == "__main__":
    renomear_arquivos()
